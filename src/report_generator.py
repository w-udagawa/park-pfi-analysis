"""Report generation: Vibrancy-first Excel (9 sheets) + Markdown output."""

import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from .vibrancy_evaluator import BADGE_KEYS, BADGE_LABELS, BADGE_MARK
from .surrounding_analysis import CATEGORIES, CATEGORY_NAMES

# ---------------------------------------------------------------------------
# Excel styling constants
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
SUB_HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
SUB_HEADER_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
LINK_FONT = Font(color="0563C1", underline="single", size=10)

RANK_FILLS = {
    "A": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "B": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "C": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "D": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

BADGE_ACTIVE_FILL = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
BADGE_INACTIVE_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

BADGE_COLORS = ["548235", "BF8F00", "2B5797", "C55A11"]  # aging/childcare/community/disaster


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_cell(cell):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center")


def _auto_width(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                val = str(cell.value or "")
                width = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, width)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _write_row(ws, row_num, values, style_func=_style_data_cell):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        style_func(cell)
    return row_num + 1


def _apply_print_settings(ws, orientation="landscape"):
    """Apply professional print settings to a worksheet."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75,
        header=0.3, footer=0.3,
    )
    ws.oddFooter.center.text = "&P / &N"


def _generate_3d_urls(lat: float, lon: float) -> Dict[str, str]:
    """Generate 3D view URLs for a given coordinate."""
    return {
        "google_earth": (
            f"https://earth.google.com/web/@{lat},{lon},100a,500d,35y,0h,60t,0r"
        ),
        "plateau_view": "https://plateauview.mlit.go.jp/",
    }


def _add_hyperlink(ws, row: int, col: int, url: str, label: str):
    """Add a clickable hyperlink to a cell."""
    cell = ws.cell(row=row, column=col, value=label)
    cell.hyperlink = url
    cell.font = LINK_FONT
    cell.border = THIN_BORDER


def _get_rank_counts(scored_parks):
    counts = {"A": 0, "B": 0, "C": 0, "D": 0}
    for p in scored_parks:
        counts[p["vibrancy"]["rank"]] += 1
    return counts


def _get_badge_counts(scored_parks):
    counts = {k: 0 for k in BADGE_KEYS}
    for p in scored_parks:
        for k in BADGE_KEYS:
            if p["badges"].get(k):
                counts[k] += 1
    return counts


def _badge_cell_value(is_active: bool, badge_key: str) -> str:
    return BADGE_MARK if is_active else ""


def _apply_badge_cell_style(cell, is_active: bool):
    cell.fill = BADGE_ACTIVE_FILL if is_active else BADGE_INACTIVE_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Sheet 1: Cover page
# ---------------------------------------------------------------------------

def _write_cover(ws, scored_parks: List[Dict], config: dict):
    ws.title = "表紙"
    municipality = config["municipality"]["name"]
    n = len(scored_parks)
    rank_counts = _get_rank_counts(scored_parks)
    badge_counts = _get_badge_counts(scored_parks)

    ws.sheet_view.showGridLines = False

    # Large title area
    ws.merge_cells("B3:G3")
    cell = ws.cell(row=3, column=2, value="Park-PFI にぎわい創出ポテンシャル分析")
    cell.font = Font(bold=True, size=22, color="2B5797")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B5:G5")
    cell = ws.cell(row=5, column=2, value=municipality)
    cell.font = Font(bold=True, size=18, color="333333")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B7:G7")
    cell = ws.cell(row=7, column=2, value=f"分析日: {datetime.now().strftime('%Y年%m月%d日')}")
    cell.font = Font(size=12, color="666666")
    cell.alignment = Alignment(horizontal="center")

    # Decorative line
    for col in range(2, 8):
        c = ws.cell(row=9, column=col)
        c.fill = HEADER_FILL
        c.border = Border(top=Side(style="medium", color="2B5797"),
                          bottom=Side(style="medium", color="2B5797"))

    # KPI section header
    row = 11
    ws.merge_cells(f"B{row}:G{row}")
    ws.cell(row=row, column=2, value="分析概要").font = Font(bold=True, size=14, color="2B5797")
    row += 2

    avg_score = sum(p["vibrancy"]["score"] for p in scored_parks) / max(n, 1)
    ab_rate = (rank_counts["A"] + rank_counts["B"]) / max(n, 1) * 100

    kpis = [
        ("対象公園数", f"{n}", ""),
        ("A/Bランク率", f"{ab_rate:.1f}%",
         f"(A: {rank_counts['A']}, B: {rank_counts['B']})"),
        ("平均にぎわいスコア", f"{avg_score:.1f}", "(0〜100)"),
        ("Top公園", scored_parks[0]["name"] if scored_parks else "—",
         f"スコア {scored_parks[0]['vibrancy']['score']}" if scored_parks else ""),
    ]

    for i, (label, value, sub) in enumerate(kpis):
        c = i * 2 + 2
        ws.cell(row=row, column=c, value=label).font = Font(size=10, color="666666")
        ws.cell(row=row + 1, column=c, value=value).font = Font(bold=True, size=16, color="2B5797")
        if sub:
            ws.cell(row=row + 2, column=c, value=sub).font = Font(size=9, color="999999")

    row += 5

    # Badge summary
    ws.merge_cells(f"B{row}:G{row}")
    ws.cell(row=row, column=2, value="サブ評価バッジ分布").font = Font(bold=True, size=12, color="2B5797")
    row += 1
    for k in BADGE_KEYS:
        count = badge_counts[k]
        pct = count / max(n, 1) * 100
        label = f"  {BADGE_LABELS[k]}: {count}公園 ({pct:.1f}%)"
        ws.cell(row=row, column=2, value=label).font = Font(size=11, color="333333")
        row += 1

    row += 1
    # Data sources summary
    ws.merge_cells(f"B{row}:G{row}")
    ws.cell(row=row, column=2, value="使用データソース").font = Font(bold=True, size=12, color="2B5797")
    row += 1
    sources = [
        "国土数値情報 都市公園データ (GeoJSON)",
        "MLIT DPF API — 鉄道駅 (S12), 福祉施設 (P14), 医療施設 (P04), 公共施設 (P02)",
        "OpenStreetMap Overpass API — 商業施設, 子育て, 教育, 高齢者施設",
    ]
    for src in sources:
        ws.cell(row=row, column=2, value=f"  {src}").font = Font(size=10, color="333333")
        row += 1

    ws.column_dimensions["A"].width = 3
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 18

    _apply_print_settings(ws)


# ---------------------------------------------------------------------------
# Sheet 2: Executive Summary
# ---------------------------------------------------------------------------

def _write_executive_summary(ws, scored_parks: List[Dict], config: dict):
    ws.title = "Executive Summary"
    n = len(scored_parks)
    rank_counts = _get_rank_counts(scored_parks)
    badge_counts = _get_badge_counts(scored_parks)

    # Title
    ws.merge_cells("A1:J1")
    cell = ws.cell(row=1, column=1, value=f"にぎわい創出ポテンシャル — {config['municipality']['name']}")
    cell.font = Font(bold=True, size=16, color="2B5797")
    cell.alignment = Alignment(horizontal="center")

    # KPI row
    row = 3
    avg_score = round(sum(p["vibrancy"]["score"] for p in scored_parks) / max(n, 1), 1)
    kpi_labels = ["対象公園数", "Aランク", "Bランク", "平均にぎわいスコア"]
    kpi_values = [
        n,
        f"{rank_counts['A']} ({rank_counts['A'] / max(n, 1) * 100:.0f}%)",
        f"{rank_counts['B']} ({rank_counts['B'] / max(n, 1) * 100:.0f}%)",
        avg_score,
    ]
    for i, (label, val) in enumerate(zip(kpi_labels, kpi_values)):
        col = i * 2 + 1
        ws.cell(row=row, column=col, value=label).font = SUB_HEADER_FONT
        ws.cell(row=row, column=col).fill = SUB_HEADER_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER
        ws.cell(row=row, column=col + 1, value=val).font = Font(bold=True, size=14)
        ws.cell(row=row, column=col + 1).border = THIN_BORDER

    # ---- Rank distribution bar chart ----
    chart_data_row = 5
    ws.cell(row=chart_data_row, column=1, value="ランク")
    ws.cell(row=chart_data_row, column=2, value="公園数")
    for i, rank in enumerate(["A", "B", "C", "D"]):
        ws.cell(row=chart_data_row + 1 + i, column=1, value=rank)
        ws.cell(row=chart_data_row + 1 + i, column=2, value=rank_counts[rank])

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "にぎわいランク分布"
    bar_chart.y_axis.title = "公園数"
    bar_chart.x_axis.title = None
    bar_chart.style = 10
    bar_chart.width = 15
    bar_chart.height = 10
    cats = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + 4)
    vals = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + 4)
    bar_chart.add_data(vals, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.shape = 4
    series = bar_chart.series[0]
    colors = ["548235", "BF8F00", "C55A11", "A5A5A5"]
    for idx, color in enumerate(colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        series.data_points.append(pt)
    ws.add_chart(bar_chart, "A10")

    # ---- Badge distribution bar chart ----
    badge_data_row = chart_data_row
    ws.cell(row=badge_data_row, column=5, value="バッジ")
    ws.cell(row=badge_data_row, column=6, value="公園数")
    for i, k in enumerate(BADGE_KEYS):
        ws.cell(row=badge_data_row + 1 + i, column=5, value=BADGE_LABELS[k])
        ws.cell(row=badge_data_row + 1 + i, column=6, value=badge_counts[k])

    badge_chart = BarChart()
    badge_chart.type = "col"
    badge_chart.title = "サブ評価バッジ分布"
    badge_chart.y_axis.title = "公園数"
    badge_chart.style = 10
    badge_chart.width = 15
    badge_chart.height = 10
    b_cats = Reference(ws, min_col=5, min_row=badge_data_row + 1, max_row=badge_data_row + 4)
    b_vals = Reference(ws, min_col=6, min_row=badge_data_row, max_row=badge_data_row + 4)
    badge_chart.add_data(b_vals, titles_from_data=True)
    badge_chart.set_categories(b_cats)
    badge_chart.shape = 4
    b_series = badge_chart.series[0]
    for idx, color in enumerate(BADGE_COLORS):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        b_series.data_points.append(pt)
    ws.add_chart(badge_chart, "E10")

    # Key findings section
    findings_row = 27
    ws.merge_cells(f"A{findings_row}:J{findings_row}")
    ws.cell(row=findings_row, column=1, value="主要知見").font = Font(bold=True, size=13, color="2B5797")
    findings_row += 1

    findings = _generate_key_findings(scored_parks, rank_counts, badge_counts, config)
    for finding in findings:
        ws.cell(row=findings_row, column=1, value=finding).font = Font(size=10)
        findings_row += 1

    _auto_width(ws)
    _apply_print_settings(ws)


def _generate_key_findings(scored_parks, rank_counts, badge_counts, config):
    """Generate key findings focused on vibrancy potential."""
    n = len(scored_parks)
    findings = []

    ab_pct = (rank_counts["A"] + rank_counts["B"]) / max(n, 1) * 100
    findings.append(
        f"1. にぎわい創出ポテンシャルが高い公園（A/Bランク）は全体の{ab_pct:.0f}%（{rank_counts['A'] + rank_counts['B']}公園）"
    )

    if scored_parks:
        top = scored_parks[0]
        active_badges = [BADGE_LABELS[k] for k in BADGE_KEYS if top["badges"].get(k)]
        badge_str = "/".join(active_badges) if active_badges else "該当バッジなし"
        findings.append(
            f"2. トップ公園: {top['name']}（にぎわいスコア: {top['vibrancy']['score']}、バッジ: {badge_str}）"
        )

    # Top 3 rank A parks
    a_parks = [p for p in scored_parks if p["vibrancy"]["rank"] == "A"][:3]
    if a_parks:
        names = "、".join(p["name"] for p in a_parks)
        findings.append(f"3. にぎわいAランク Top3: {names}")

    # Badge cross-analysis: vibrancy + childcare (multi-use opportunity)
    multi = [
        p for p in scored_parks
        if p["vibrancy"]["rank"] in ("A", "B") and p["badges"].get("childcare")
    ]
    findings.append(
        f"4. にぎわい＋子育て両立候補（A/Bランク かつ 子育てバッジ）: {len(multi)}公園"
    )

    # Disaster badge with vibrancy
    disaster_vibrant = [
        p for p in scored_parks
        if p["badges"].get("disaster") and p["vibrancy"]["score"] >= config["ranks"]["C"]
    ]
    findings.append(
        f"5. 防災＋にぎわい複合型候補（防災バッジ かつ にぎわいC以上）: {len(disaster_vibrant)}公園"
    )

    # Aging + vibrancy = renewal opportunity
    aging_vibrant = [
        p for p in scored_parks
        if p["badges"].get("aging") and p["vibrancy"]["rank"] in ("A", "B")
    ]
    findings.append(
        f"6. 老朽＋にぎわい＝再生型候補（老朽バッジ かつ にぎわいA/B）: {len(aging_vibrant)}公園"
    )

    return findings


# ---------------------------------------------------------------------------
# Sheet 3: Methodology
# ---------------------------------------------------------------------------

def _write_methodology(ws, config: dict):
    ws.title = "分析手法"
    ws.sheet_view.showGridLines = False

    def section_title(row: int, text: str) -> int:
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=13, color="2B5797")
        return row + 1

    def paragraph(row: int, text: str, *, indent: bool = False, bold: bool = False) -> int:
        prefix = "  " if indent else ""
        cell = ws.cell(row=row, column=1, value=prefix + text)
        cell.font = Font(size=11, bold=bold, color="333333")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"A{row}:F{row}")
        ws.row_dimensions[row].height = max(18, 16 * (1 + text.count("。") // 2))
        return row + 1

    def blank(row: int, n: int = 1) -> int:
        return row + n

    def styled_header_row(row: int, headers: List[str]) -> int:
        _write_row(ws, row, headers)
        for cell in ws[row]:
            cell.fill = SUB_HEADER_FILL
            cell.font = SUB_HEADER_FONT
            cell.border = THIN_BORDER
        return row + 1

    row = 1
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="分析手法・評価基準").font = Font(bold=True, size=16, color="2B5797")
    row += 2

    v = config["vibrancy"]
    b = config["badges"]

    # -----------------------------------------------------------------
    # A. 評価軸選定の背景
    # -----------------------------------------------------------------
    row = section_title(row, "A. 評価軸選定の背景")
    row = paragraph(row, "本分析は「Park-PFI（公募設置管理制度）導入ポテンシャル」を定量化するが、その軸を「にぎわい創出ポテンシャル」単一に絞っている。その背景は下記の通り。")
    row = blank(row)
    row = paragraph(row, "1. 北谷公園（渋谷区、約960m²、Park-PFI 実施事例）の現地調査で、事業者・利用者インタビューから「人流の量」と「人流の属性（周辺建物用途）」が Park-PFI の成否を決定する最大要因であることが確認された。", indent=True)
    row = paragraph(row, "2. 従来版（5類型分類＋5軸複合スコア）では類型間のスコア競合により、本来にぎわい創出型として有望な公園が「老朽再生型」「子育て支援型」に分類され埋もれていた（世田谷区152公園中、にぎわい創出型が Primary 判定となったのは1公園のみ）。", indent=True)
    row = paragraph(row, "3. そこで、にぎわい創出ポテンシャルをメイン評価軸とし、「老朽化」「子育て周辺」「地域拠点性」「防災ポテンシャル」はサブ評価（●バッジ）として併記する構造に変更した。これにより、メイン軸での比較可能性を担保しつつ、導入検討時に必要な文脈情報を失わない。", indent=True)
    row = blank(row)

    # -----------------------------------------------------------------
    # B. にぎわいスコアの計算式と根拠
    # -----------------------------------------------------------------
    row = section_title(row, "B. にぎわいスコアの計算式と根拠")
    row = paragraph(row, "計算式:", bold=True)
    row = paragraph(row, f"  vibrancy_score = flow_percentile × {v['flow_weight']} + commercial_score × {v['commercial_weight']}")
    row = paragraph(row, f"  commercial_score = min(commercial_count / {v['commercial_divisor']} × 100, 100)")
    row = blank(row)
    row = paragraph(row, "重み配分の根拠:", bold=True)
    row = paragraph(row, f"・flow_percentile × {v['flow_weight']}: 「人が通らない場所に施設を作っても稼働しない」という事業者サイドの最大制約を反映。人流は Park-PFI の必要条件であり、最重要項目として60%を配分。", indent=True)
    row = paragraph(row, f"・commercial_score × {v['commercial_weight']}: 「目的意識を持って訪れる人流」の補完指標。周辺に飲食・物販施設が集積しているエリアでは、公園利用が食事・買い物の延長に組み込まれやすく、滞留時間と来訪頻度が高まる。", indent=True)
    row = blank(row)
    row = paragraph(row, f"commercial_divisor = {v['commercial_divisor']} の根拠:", bold=True)
    row = paragraph(row, f"世田谷区の商業集積上位エリア（三軒茶屋・下北沢・二子玉川周辺）における、半径{config['facility']['radius']}m圏内の目的地型商業数の水準（約{v['commercial_divisor']}件前後）を満点（commercial_score=100）とするスケール設計。他自治体展開時は config.vibrancy.commercial_divisor で調整可能。", indent=True)
    row = blank(row)
    row = paragraph(row, "ハード足切りを廃止した理由:", bold=True)
    row = paragraph(row, "旧実装では「流量上位40%（percentile≥60）」かつ「商業施設2件以上」のAND条件を足切りとしていたため、約90%の公園がスコア0となっていた。相対評価として連続スコアに変更することで、低スコア帯の公園間でも比較が可能になり、サブバッジと組み合わせた多面的な候補抽出ができるようになった。", indent=True)
    row = blank(row)
    row = paragraph(row, "パーセンタイル正規化の意味:", bold=True)
    row = paragraph(row, "flow_percentile は駅乗降客数の距離減衰合計を自治体内で順位化した相対指標。絶対的な乗降客数ではなく「その自治体における人流の多さランク」を表すため、都心部でも地方都市でも同じ物差しで評価できる設計としている。", indent=True)
    row = blank(row)
    row = paragraph(row, f"pedestrian_flow の距離減衰式:", bold=True)
    row = paragraph(row, f"raw_score = Σ(駅乗降客数 × exp(-距離 / {config['flow']['decay_constant']}m))（半径{config['flow']['max_radius']}m以内の全駅）→ 全公園内でパーセンタイル正規化し 0–100 に変換。", indent=True)
    row = blank(row)

    # -----------------------------------------------------------------
    # C. 「目的地型商業」の定義
    # -----------------------------------------------------------------
    row = section_title(row, "C. 「目的地型商業」の定義")
    row = paragraph(row, "commercial_count では、以下の区分で OSM の商業施設を絞り込んでいる。", indent=False)
    row = blank(row)
    row = paragraph(row, "対象（目的地型）:", bold=True)
    row = paragraph(row, "レストラン・カフェ・バー・居酒屋、書店、衣料品店、電化製品店、雑貨・ギフト、美容・エンタメ施設、ベーカリー等、「その場所に行くこと自体が目的となる」業態。", indent=True)
    row = blank(row)
    row = paragraph(row, "除外（日常型）:", bold=True)
    row = paragraph(row, "コンビニ・ファストフード・ドラッグストア・スーパーマーケット等、生活動線上の「ついで利用」が中心の業態。", indent=True)
    row = blank(row)
    row = paragraph(row, "除外の根拠:", bold=True)
    row = paragraph(row, "日常型商業は来訪の目的ではなく経路上の存在であり、公園に立ち寄る動機にならない。南池袋公園・北谷公園等の成功事例では、周辺の目的地型商業と公園が「訪問動線上に組み込まれる」構造が共通しており、この層の密度が Park-PFI 事業の滞留時間・回遊性に直結する。", indent=True)
    row = blank(row)
    row = paragraph(row, "データソース:", bold=True)
    row = paragraph(row, "OpenStreetMap Overpass API（src/data_collector.py の DESTINATION_SHOP_TYPES で分類。世田谷区では約43%の商業タグが日常型として除外された）。", indent=True)
    row = blank(row)

    # -----------------------------------------------------------------
    # D. ランク閾値の根拠
    # -----------------------------------------------------------------
    row = section_title(row, "D. ランク閾値の意味")
    row = styled_header_row(row, ["ランク", "閾値", "解釈"])
    rank_rows = [
        ("A", f"スコア ≥ {config['ranks']['A']}",
         "即事業化検討レベル。人流・商業ともに非常に高水準で、Park-PFI 単独でも収益性が見込める"),
        ("B", f"スコア ≥ {config['ranks']['B']}",
         "条件次第で有望。導入事業者との個別協議・サウンディング調査を推奨"),
        ("C", f"スコア ≥ {config['ranks']['C']}",
         "単独では難。サブバッジとの組合せで活用検討（例: 防災+にぎわいC → 平時にぎわい/災害時拠点の複合スキーム）"),
        ("D", f"スコア < {config['ranks']['C']}",
         "Park-PFI 単独事業は困難。指定管理・PFS 等の他手法や、近隣施設との連携事業を検討"),
    ]
    for rank, threshold, meaning in rank_rows:
        ws.cell(row=row, column=1, value=rank).border = THIN_BORDER
        ws.cell(row=row, column=1).fill = RANK_FILLS[rank]
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=2, value=threshold).border = THIN_BORDER
        mcell = ws.cell(row=row, column=3, value=meaning)
        mcell.border = THIN_BORDER
        mcell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 32
        row += 1
    row = blank(row)

    # -----------------------------------------------------------------
    # E. サブ評価バッジの設計根拠
    # -----------------------------------------------------------------
    row = section_title(row, "E. サブ評価バッジの設計根拠")
    row = paragraph(row, "バッジはメイン軸（にぎわいスコア）を単一化する一方で、Park-PFI 導入検討時に必要な文脈情報（築年数・周辺施設構成・面積）を失わないためのサブ評価。該当すれば ● 、非該当なら空白で表示する。")
    row = blank(row)
    row = styled_header_row(row, ["バッジ", "判定基準", "Park-PFI 活用シナリオ"])

    badge_info = [
        (BADGE_LABELS["aging"],
         f"供用開始から{b['aging_years']}年以上経過",
         "老朽施設の更新需要と Park-PFI を組み合わせ、自治体の財政負担を軽減。再整備費用を事業者側設備投資で賄えるスキームが成立しやすい"),
        (BADGE_LABELS["childcare"],
         f"半径{config['facility']['radius']}m以内に子育て・教育施設が{b['childcare_min']}件以上",
         "子育て世帯の日常利用動線に組み込み、親子向けカフェ・屋内プレイスペース・体験型施設の導入候補。送迎・放課後利用のニーズが見込める"),
        (BADGE_LABELS["community_hub"],
         f"公共施設{b['community_public_min']}件以上 かつ 医療施設{b['community_medical_min']}件以上近接",
         "高齢者・地域住民の多世代利用前提の交流施設・集会所・健康増進拠点との連携事業。行政窓口サテライトや健康相談との組合せも可能"),
        (BADGE_LABELS["disaster"],
         f"面積{b['disaster_min_area']:,}m²以上、近隣・地区・総合公園",
         "平時はにぎわい創出、災害時は避難地・防災拠点となる複合 PFI スキーム候補。備蓄倉庫付カフェ、発電設備付東屋等のハイブリッド設計"),
    ]
    for name, criteria, scenario in badge_info:
        name_cell = ws.cell(row=row, column=1, value=name)
        name_cell.border = THIN_BORDER
        name_cell.font = Font(bold=True)
        name_cell.alignment = Alignment(horizontal="center", vertical="center")
        cr_cell = ws.cell(row=row, column=2, value=criteria)
        cr_cell.border = THIN_BORDER
        cr_cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        sc_cell = ws.cell(row=row, column=4, value=scenario)
        sc_cell.border = THIN_BORDER
        sc_cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 48
        row += 1
    row = blank(row)

    # -----------------------------------------------------------------
    # F. データソース一覧
    # -----------------------------------------------------------------
    row = section_title(row, "F. データソース一覧")
    row = styled_header_row(row, ["データソース", "用途", "取得方法"])
    data_sources = [
        ("国土数値情報 都市公園データ", "公園の位置・面積・種別・供用開始年", "GeoJSON"),
        ("MLIT DPF — 鉄道駅 (S12)", "駅位置・乗降客数（2023年実績）", "GraphQL API"),
        ("MLIT DPF — 福祉施設 (P14)", "子育て・高齢者施設", "GraphQL API"),
        ("MLIT DPF — 医療施設 (P04)", "病院・診療所", "GraphQL API"),
        ("MLIT DPF — 公共施設 (P02)", "公民館・図書館等", "GraphQL API"),
        ("OpenStreetMap (Overpass API)", "商業施設・飲食店・教育施設", "Overpass QL"),
    ]
    for source, usage, method in data_sources:
        ws.cell(row=row, column=1, value=source).border = THIN_BORDER
        ws.cell(row=row, column=2, value=usage).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=5, value=method).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        row += 1
    row = blank(row)

    # -----------------------------------------------------------------
    # G. 限界と注意点
    # -----------------------------------------------------------------
    row = section_title(row, "G. 限界と注意点")
    cautions = [
        "OSM の商業データは地域ごとのメンテナンス頻度に差があり、地方都市等ではカバレッジ不足により commercial_count が過小評価となる可能性がある。",
        "駅乗降客数は2023年実績のため、再開発・鉄道新線開業等による最新の人流変化は未反映。",
        f"対象面積下限 {config.get('min_area', 960)}m² は北谷公園（約960m²）の事業実績を参考にした目安。これ未満の小規模公園は分析対象外。",
        "バッジしきい値（築30年、子育て2件、公共1件かつ医療1件、防災3,000m²）は世田谷区向けの標準設定。人口密度・都市構造の異なる自治体では config/badges で調整推奨。",
        "目的地型／日常型商業の分類は OSM タグに対する機械的ルールベース判定。個店レベルでは誤分類の可能性があり、エリア特性に応じた再調整が必要な場合がある。",
        "本分析は机上スクリーニングであり、実際の事業化判断にあたっては、現地調査・地元合意形成・サウンディング調査等を並行して実施すること。",
    ]
    for caution in cautions:
        row = paragraph(row, f"・{caution}", indent=True)

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    _apply_print_settings(ws)


# ---------------------------------------------------------------------------
# Sheet 4: Vibrancy Ranking (main)
# ---------------------------------------------------------------------------

def _write_vibrancy_ranking(ws, scored_parks: List[Dict], config: dict):
    ws.title = "にぎわいランキング"
    enable_3d = config["report"].get("enable_3d_links", True)

    # Title
    ws.merge_cells("A1:P1")
    cell = ws.cell(row=1, column=1, value=f"にぎわい創出ポテンシャル 全公園ランキング ({len(scored_parks)}公園)")
    cell.font = Font(bold=True, size=14, color="2B5797")
    cell.alignment = Alignment(horizontal="center")

    headers = [
        "順位", "公園名", "にぎわいスコア", "ランク",
        "流量%ile", "商業数",
        "老朽", "子育て", "地域拠点", "防災",
        "面積(m²)", "種別", "開園年",
    ]
    if enable_3d:
        headers.append("3Dビュー")

    _write_row(ws, 3, headers)
    _style_header(ws, 3)
    header_row = 3

    for i, p in enumerate(scored_parks):
        v = p["vibrancy"]
        row = header_row + 1 + i
        vals = [
            p["rank_position"],
            p["name"],
            v["score"],
            v["rank"],
            v["flow_percentile"],
            v["commercial_count"],
            _badge_cell_value(p["badges"]["aging"], "aging"),
            _badge_cell_value(p["badges"]["childcare"], "childcare"),
            _badge_cell_value(p["badges"]["community_hub"], "community_hub"),
            _badge_cell_value(p["badges"]["disaster"], "disaster"),
            p["area_m2"],
            p["type_name"],
            p.get("year_opened", ""),
        ]
        _write_row(ws, row, vals)

        # Rank fill
        ws.cell(row=row, column=4).fill = RANK_FILLS.get(v["rank"], PatternFill())
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

        # Badge cell styling
        for bi, bk in enumerate(BADGE_KEYS):
            _apply_badge_cell_style(ws.cell(row=row, column=7 + bi), p["badges"][bk])

        # Number formats
        ws.cell(row=row, column=3).number_format = '0.0'
        ws.cell(row=row, column=5).number_format = '0.0'
        ws.cell(row=row, column=11).number_format = '#,##0'

        # 3D link
        if enable_3d and p.get("lat") and p.get("lon"):
            urls = _generate_3d_urls(p["lat"], p["lon"])
            _add_hyperlink(ws, row, 14, urls["google_earth"], "3D")

    last_row = header_row + len(scored_parks)

    # Color scale for vibrancy score (column C)
    ws.conditional_formatting.add(
        f"C{header_row + 1}:C{last_row}",
        ColorScaleRule(
            start_type="min", start_color="FFC7CE",
            mid_type="percentile", mid_value=50, mid_color="FFEB9C",
            end_type="max", end_color="C6EFCE",
        ),
    )
    # Data bar for flow percentile (E)
    ws.conditional_formatting.add(
        f"E{header_row + 1}:E{last_row}",
        DataBarRule(start_type="min", end_type="max", color="5B9BD5"),
    )
    # Data bar for commercial count (F)
    ws.conditional_formatting.add(
        f"F{header_row + 1}:F{last_row}",
        DataBarRule(start_type="min", end_type="max", color="ED7D31"),
    )

    # Freeze: header + park name column
    ws.freeze_panes = f"C{header_row + 1}"

    _apply_print_settings(ws)
    ws.print_title_rows = f"{header_row}:{header_row}"
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 5: Pedestrian flow
# ---------------------------------------------------------------------------

def _write_pedestrian_flow(ws, scored_parks: List[Dict]):
    ws.title = "歩行者流量"

    headers = [
        "順位", "公園名", "流量スコア(正規化)", "流量スコア(生値)",
        "最寄駅", "最寄駅距離(m)", "最寄駅乗降客数",
        "800m内駅数",
    ]
    _write_row(ws, 1, headers)
    _style_header(ws, 1)

    by_flow = sorted(scored_parks, key=lambda p: -p["flow"].get("normalized_score", 0))

    for i, p in enumerate(by_flow):
        flow = p.get("flow", {})
        nearest = flow.get("nearest_station")
        row = i + 2
        vals = [
            i + 1,
            p["name"],
            round(flow.get("normalized_score", 0), 1),
            round(flow.get("raw_score", 0), 1),
            nearest["name"] if nearest else "なし",
            round(nearest["distance_m"]) if nearest else "",
            nearest["ridership"] if nearest else "",
            flow.get("station_count", 0),
        ]
        _write_row(ws, row, vals)
        ws.cell(row=row, column=3).number_format = '0.0'
        ws.cell(row=row, column=4).number_format = '0.0'
        if nearest and nearest.get("ridership"):
            ws.cell(row=row, column=7).number_format = '#,##0'

    last_row = len(by_flow) + 1

    ws.conditional_formatting.add(
        f"C2:C{last_row}",
        DataBarRule(start_type="min", end_type="max", color="5B9BD5"),
    )
    ws.conditional_formatting.add(
        f"G2:G{last_row}",
        ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max", end_color="4472C4",
        ),
    )

    ws.freeze_panes = "C2"
    _auto_width(ws)
    _apply_print_settings(ws)
    ws.print_title_rows = "1:1"


# ---------------------------------------------------------------------------
# Sheet 6: Surrounding facilities
# ---------------------------------------------------------------------------

def _write_surrounding_facilities(ws, scored_parks: List[Dict]):
    ws.title = "周辺施設"

    cat_headers = [CATEGORY_NAMES[c] for c in CATEGORIES]
    headers = ["順位", "公園名"] + cat_headers + ["合計", "カテゴリ多様性"]
    _write_row(ws, 1, headers)
    _style_header(ws, 1)

    by_total = sorted(
        scored_parks,
        key=lambda p: -p.get("surroundings", {}).get("total_count", 0),
    )

    for i, p in enumerate(by_total):
        surr = p.get("surroundings", {})
        cat_counts = [surr.get(c, {}).get("count", 0) for c in CATEGORIES]
        row = i + 2
        vals = [
            i + 1,
            p["name"],
        ] + cat_counts + [
            surr.get("total_count", 0),
            f"{surr.get('diversity', 0)}/{len(CATEGORIES)}",
        ]
        _write_row(ws, row, vals)

    last_row = len(by_total) + 1
    n_cats = len(CATEGORIES)

    for idx in range(n_cats):
        col_letter = get_column_letter(3 + idx)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{last_row}",
            ColorScaleRule(
                start_type="num", start_value=0, start_color="FFFFFF",
                end_type="max", end_color="2B5797",
            ),
        )

    total_col = get_column_letter(3 + n_cats)
    ws.conditional_formatting.add(
        f"{total_col}2:{total_col}{last_row}",
        DataBarRule(start_type="min", end_type="max", color="ED7D31"),
    )

    ws.freeze_panes = "C2"
    _auto_width(ws)
    _apply_print_settings(ws)
    ws.print_title_rows = "1:1"


# ---------------------------------------------------------------------------
# Sheet 7: Park carte (Top N profiles, badges + score breakdown)
# ---------------------------------------------------------------------------

def _write_park_carte(ws, scored_parks: List[Dict], config: dict):
    ws.title = "公園カルテ"
    top_n = config["report"]["top_n_carte"]
    enable_3d = config["report"].get("enable_3d_links", True)

    row = 1

    for p in scored_parks[:top_n]:
        v = p["vibrancy"]

        # Park header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=f"#{p['rank_position']} {p['name']}")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        for col in range(2, 7):
            ws.cell(row=row, column=col).fill = HEADER_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

        # Basic info
        active_badges = [BADGE_LABELS[k] for k in BADGE_KEYS if p["badges"].get(k)]
        badge_str = "・".join(active_badges) if active_badges else "該当バッジなし"

        info = [
            ("にぎわいランク", v["rank"]),
            ("にぎわいスコア", v["score"]),
            ("バッジ", badge_str),
            ("面積(m²)", f"{p['area_m2']:,}"),
            ("種別", p["type_name"]),
            ("供用開始年", p.get("year_opened", "不明")),
        ]

        if enable_3d and p.get("lat") and p.get("lon"):
            urls = _generate_3d_urls(p["lat"], p["lon"])
        else:
            urls = None

        for label, val in info:
            ws.cell(row=row, column=1, value=label).font = SUB_HEADER_FONT
            ws.cell(row=row, column=1).fill = SUB_HEADER_FILL
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=val).border = THIN_BORDER
            if label == "にぎわいランク":
                ws.cell(row=row, column=2).fill = RANK_FILLS.get(v["rank"], PatternFill())
            row += 1

        if urls:
            ws.cell(row=row, column=1, value="3Dビュー").font = SUB_HEADER_FONT
            ws.cell(row=row, column=1).fill = SUB_HEADER_FILL
            ws.cell(row=row, column=1).border = THIN_BORDER
            _add_hyperlink(ws, row, 2, urls["google_earth"], "Google Earth 3D")
            _add_hyperlink(ws, row, 3, urls["plateau_view"], "PLATEAU VIEW")
            row += 1

        # Vibrancy score breakdown
        ws.cell(row=row, column=1, value="にぎわいスコア内訳").font = SUB_HEADER_FONT
        row += 1

        breakdown = [
            (f"流量%ile × {config['vibrancy']['flow_weight']}",
             round(v["flow_percentile"] * config["vibrancy"]["flow_weight"], 1),
             f"(flow_percentile = {v['flow_percentile']})"),
            (f"商業スコア × {config['vibrancy']['commercial_weight']}",
             round(v["commercial_score"] * config["vibrancy"]["commercial_weight"], 1),
             f"(commercial_count = {v['commercial_count']}, score = {v['commercial_score']})"),
            ("合計", v["score"], ""),
        ]
        for label, val, note in breakdown:
            ws.cell(row=row, column=1, value=label).border = THIN_BORDER
            ws.cell(row=row, column=2, value=val).border = THIN_BORDER
            ws.cell(row=row, column=2).number_format = '0.0'
            if note:
                ws.cell(row=row, column=3, value=note).border = THIN_BORDER
                ws.cell(row=row, column=3).font = Font(size=9, color="666666")
            row += 1

        # Nearby stations
        flow = p.get("flow", {})
        stations = flow.get("station_details", [])
        if stations:
            ws.cell(row=row, column=1, value="最寄駅").font = SUB_HEADER_FONT
            row += 1
            for st in stations[:5]:
                ws.cell(row=row, column=1, value=st["name"]).border = THIN_BORDER
                ws.cell(row=row, column=2, value=f"{st['distance_m']}m").border = THIN_BORDER
                ws.cell(row=row, column=3, value=f"乗降客数: {st['ridership']:,}").border = THIN_BORDER
                row += 1

        # Surrounding facilities summary
        surr = p.get("surroundings", {})
        ws.cell(row=row, column=1, value="周辺施設").font = SUB_HEADER_FONT
        row += 1
        for cat in CATEGORIES:
            cat_data = surr.get(cat, {})
            ws.cell(row=row, column=1, value=CATEGORY_NAMES[cat]).border = THIN_BORDER
            ws.cell(row=row, column=2, value=f"{cat_data.get('count', 0)}施設").border = THIN_BORDER
            nearest = cat_data.get("nearest")
            if nearest:
                ws.cell(row=row, column=3, value=f"最寄: {nearest['name']} ({nearest['distance_m']}m)").border = THIN_BORDER
            row += 1

        row += 2  # blank rows between parks

    _auto_width(ws)
    _apply_print_settings(ws)


# ---------------------------------------------------------------------------
# Sheet 8: Badge analysis
# ---------------------------------------------------------------------------

def _write_badge_analysis(ws, scored_parks: List[Dict]):
    ws.title = "バッジ別分析"

    row = 1
    for bk in BADGE_KEYS:
        tagged = [p for p in scored_parks if p["badges"].get(bk)]

        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(
            row=row, column=1,
            value=f"{BADGE_LABELS[bk]} ({len(tagged)}公園)",
        )
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, 7):
            ws.cell(row=row, column=col).fill = HEADER_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

        if not tagged:
            ws.cell(row=row, column=1, value="該当公園なし").font = Font(color="999999", italic=True)
            row += 2
            continue

        # Average vibrancy score for this badge group
        avg = sum(p["vibrancy"]["score"] for p in tagged) / len(tagged)
        ws.cell(row=row, column=1, value=f"  平均にぎわいスコア: {avg:.1f}").font = Font(italic=True, color="666666")
        row += 1

        headers = ["順位(全体)", "公園名", "にぎわいスコア", "ランク", "面積(m²)", "開園年"]
        _write_row(ws, row, headers)
        for cell in ws[row]:
            cell.fill = SUB_HEADER_FILL
            cell.font = SUB_HEADER_FONT
            cell.border = THIN_BORDER
        row += 1

        tagged.sort(key=lambda p: -p["vibrancy"]["score"])
        for p in tagged:
            v = p["vibrancy"]
            vals = [
                p["rank_position"],
                p["name"],
                v["score"],
                v["rank"],
                p["area_m2"],
                p.get("year_opened", ""),
            ]
            _write_row(ws, row, vals)
            ws.cell(row=row, column=3).number_format = '0.0'
            ws.cell(row=row, column=4).fill = RANK_FILLS.get(v["rank"], PatternFill())
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5).number_format = '#,##0'
            row += 1

        row += 1

    _auto_width(ws)
    _apply_print_settings(ws)


# ---------------------------------------------------------------------------
# Sheet 9: Data quality
# ---------------------------------------------------------------------------

def _write_data_quality(ws, scored_parks: List[Dict], validation: Optional[Dict] = None):
    ws.title = "データ品質"

    start_row = 1
    if validation:
        ws.merge_cells("A1:L1")
        ws.cell(row=1, column=1, value="データカバレッジ診断").font = Font(bold=True, size=14, color="2B5797")

        row = 3
        summary = validation.get("summary", "N/A")
        status_colors = {
            "PASS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "WARN": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "FAIL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        }
        ws.cell(row=row, column=1, value="総合判定").font = SUB_HEADER_FONT
        ws.cell(row=row, column=1).fill = SUB_HEADER_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        status_cell = ws.cell(row=row, column=2, value=summary)
        status_cell.font = Font(bold=True, size=12)
        status_cell.fill = status_colors.get(summary, PatternFill())
        status_cell.border = THIN_BORDER

        ws.cell(row=row, column=3, value="総合スコア").font = SUB_HEADER_FONT
        ws.cell(row=row, column=3).fill = SUB_HEADER_FILL
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f"{validation.get('overall_score', 0)}").border = THIN_BORDER
        row += 2

        ws.cell(row=row, column=1, value="データ件数").font = SUB_HEADER_FONT
        row += 1
        counts = validation.get("counts", {})
        _write_row(ws, row, ["データ種別", "件数"])
        for cell in ws[row]:
            cell.fill = SUB_HEADER_FILL
            cell.font = SUB_HEADER_FONT
            cell.border = THIN_BORDER
        row += 1

        count_labels = {
            "parks": "公園",
            "stations": "駅",
            "welfare": "福祉施設",
            "medical": "医療施設",
            "public": "公共施設",
            "commercial": "商業施設 (OSM)",
        }
        for key, label in count_labels.items():
            row = _write_row(ws, row, [label, counts.get(key, 0)])

        row += 1
        ws.cell(row=row, column=1, value="カバレッジスコア").font = SUB_HEADER_FONT
        row += 1
        _write_row(ws, row, ["カテゴリ", "スコア"])
        for cell in ws[row]:
            cell.fill = SUB_HEADER_FILL
            cell.font = SUB_HEADER_FONT
            cell.border = THIN_BORDER
        row += 1

        score_labels = {
            "parks": "公園データ",
            "stations": "駅データ",
            "welfare": "福祉施設",
            "medical": "医療施設",
            "public": "公共施設",
            "commercial": "商業データ",
        }
        for key, label in score_labels.items():
            score_val = validation.get("coverage_scores", {}).get(key, "N/A")
            row = _write_row(ws, row, [label, score_val])

        issues = validation.get("issues", [])
        if issues:
            row += 1
            ws.cell(row=row, column=1, value="検出された問題").font = SUB_HEADER_FONT
            row += 1
            _write_row(ws, row, ["重要度", "カテゴリ", "内容", "詳細"])
            for cell in ws[row]:
                cell.fill = SUB_HEADER_FILL
                cell.font = SUB_HEADER_FONT
                cell.border = THIN_BORDER
            row += 1

            severity_fills = {
                "HIGH": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                "LOW": PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
            }
            for issue in issues:
                _write_row(ws, row, [
                    issue["severity"],
                    issue["category"],
                    issue["message"],
                    issue["detail"],
                ])
                ws.cell(row=row, column=1).fill = severity_fills.get(issue["severity"], PatternFill())
                row += 1

        row += 2
        start_row = row

    # Per-park data quality table
    ws.cell(row=start_row, column=1, value="公園別データ完備率").font = Font(bold=True, size=14, color="2B5797")
    start_row += 1

    headers = [
        "公園名", "座標", "面積", "供用開始年", "駅データ",
        "商業施設", "子育て", "教育", "高齢者", "公共", "医療",
        "完備率",
    ]
    _write_row(ws, start_row, headers)
    _style_header(ws, start_row)
    data_start = start_row + 1

    for i, p in enumerate(sorted(scored_parks, key=lambda x: x["name"])):
        has_coords = "○" if p["lat"] and p["lon"] else "×"
        has_area = "○" if p["area_m2"] > 0 else "×"
        has_year = "○" if p.get("year_opened") and p["year_opened"] > 0 else "×"
        has_stations = "○" if p.get("flow", {}).get("station_count", 0) > 0 else "△"

        surr = p.get("surroundings", {})
        cat_flags = []
        for cat in CATEGORIES:
            count = surr.get(cat, {}).get("count", 0)
            cat_flags.append("○" if count > 0 else "△")

        all_flags = [has_coords, has_area, has_year, has_stations] + cat_flags
        complete = sum(1 for f in all_flags if f == "○")
        total_checks = len(all_flags)
        rate = f"{complete / total_checks * 100:.0f}%"

        row = data_start + i
        vals = [p["name"], has_coords, has_area, has_year, has_stations] + cat_flags + [rate]
        _write_row(ws, row, vals)

    ws.freeze_panes = f"B{data_start}"
    _auto_width(ws)
    _apply_print_settings(ws)
    ws.print_title_rows = f"{start_row}:{start_row}"


# ---------------------------------------------------------------------------
# Main Excel generator
# ---------------------------------------------------------------------------

def generate_excel(
    scored_parks: List[Dict],
    config: dict,
    output_dir: str,
    validation: Optional[Dict] = None,
) -> str:
    """Generate vibrancy-first Excel report (9 sheets). Returns output file path."""
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()

    # Sheet 1: Cover
    ws = wb.active
    _write_cover(ws, scored_parks, config)

    # Sheet 2: Executive Summary
    _write_executive_summary(wb.create_sheet(), scored_parks, config)

    # Sheet 3: Methodology
    _write_methodology(wb.create_sheet(), config)

    # Sheet 4: Vibrancy Ranking (main)
    _write_vibrancy_ranking(wb.create_sheet(), scored_parks, config)

    # Sheet 5: Pedestrian flow
    _write_pedestrian_flow(wb.create_sheet(), scored_parks)

    # Sheet 6: Surrounding facilities
    _write_surrounding_facilities(wb.create_sheet(), scored_parks)

    # Sheet 7: Park carte (Top N)
    _write_park_carte(wb.create_sheet(), scored_parks, config)

    # Sheet 8: Badge analysis
    _write_badge_analysis(wb.create_sheet(), scored_parks)

    # Sheet 9: Data quality
    _write_data_quality(wb.create_sheet(), scored_parks, validation)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    municipality = config["municipality"]["name"]
    filename = f"park_pfi_analysis_{municipality}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"  Excel saved: {filepath}")
    return filepath


# ---------------------------------------------------------------------------
# Markdown report
# ---------------------------------------------------------------------------

def generate_markdown(
    scored_parks: List[Dict],
    config: dict,
    output_dir: str,
    validation: Optional[Dict] = None,
) -> str:
    """Generate Markdown summary report (vibrancy-first)."""
    os.makedirs(output_dir, exist_ok=True)
    municipality = config["municipality"]["name"]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"park_pfi_analysis_{municipality}_{timestamp}.md"
    filepath = os.path.join(output_dir, filename)
    enable_3d = config["report"].get("enable_3d_links", True)
    n = len(scored_parks)

    lines = []
    lines.append(f"# Park-PFI にぎわい創出ポテンシャル分析 — {municipality}")
    lines.append("")
    lines.append(f"分析日: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"対象公園数: {n}")
    lines.append("")

    # Rank distribution
    lines.append("## にぎわいランク分布")
    lines.append("")
    rank_counts = _get_rank_counts(scored_parks)
    lines.append("| ランク | 公園数 | 割合 |")
    lines.append("|--------|--------|------|")
    for rank in ["A", "B", "C", "D"]:
        count = rank_counts[rank]
        pct = f"{count / max(n, 1) * 100:.1f}%"
        lines.append(f"| {rank} | {count} | {pct} |")
    lines.append("")

    # Badge distribution
    lines.append("## サブ評価バッジ分布")
    lines.append("")
    badge_counts = _get_badge_counts(scored_parks)
    lines.append("| バッジ | 公園数 | 割合 |")
    lines.append("|--------|--------|------|")
    for k in BADGE_KEYS:
        count = badge_counts[k]
        pct = f"{count / max(n, 1) * 100:.1f}%"
        lines.append(f"| {BADGE_LABELS[k]} | {count} | {pct} |")
    lines.append("")

    # Top N parks
    top_n = config["report"]["top_n_summary"]
    lines.append(f"## にぎわいスコア Top {top_n}")
    lines.append("")

    if enable_3d:
        lines.append("| 順位 | 公園名 | ランク | にぎわい | 老朽 | 子育て | 地域拠点 | 防災 | 面積(m²) | 3Dビュー |")
        lines.append("|------|--------|--------|----------|------|--------|----------|------|----------|----------|")
    else:
        lines.append("| 順位 | 公園名 | ランク | にぎわい | 老朽 | 子育て | 地域拠点 | 防災 | 面積(m²) |")
        lines.append("|------|--------|--------|----------|------|--------|----------|------|----------|")

    for p in scored_parks[:top_n]:
        v = p["vibrancy"]
        mark = lambda k: BADGE_MARK if p["badges"].get(k) else ""
        base = (
            f"| {p['rank_position']} | {p['name']} | {v['rank']} | {v['score']} | "
            f"{mark('aging')} | {mark('childcare')} | {mark('community_hub')} | {mark('disaster')} | "
            f"{p['area_m2']:,}"
        )
        if enable_3d and p.get("lat") and p.get("lon"):
            urls = _generate_3d_urls(p["lat"], p["lon"])
            base += f" | [Earth]({urls['google_earth']}) / [PLATEAU]({urls['plateau_view']})"
        elif enable_3d:
            base += " | — "
        base += " |"
        lines.append(base)
    lines.append("")

    # Data coverage
    if validation:
        lines.append("## データカバレッジ診断")
        lines.append("")
        lines.append(f"**総合判定: {validation.get('summary', 'N/A')}** (スコア: {validation.get('overall_score', 0)})")
        lines.append("")
        issues = validation.get("issues", [])
        if issues:
            lines.append("| 重要度 | カテゴリ | 内容 |")
            lines.append("|--------|----------|------|")
            for issue in issues:
                lines.append(f"| {issue['severity']} | {issue['category']} | {issue['message']} |")
            lines.append("")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"  Markdown saved: {filepath}")
    return filepath
